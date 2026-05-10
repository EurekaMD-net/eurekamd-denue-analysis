#!/usr/bin/env python3
r"""
Convert CNBV Panorama Anual de Inclusión Financiera 2025 XLSX to a clean CSV.

Source: Anexo_Panorama_2025.xlsx (operator-supplied via Drive 2026-05-10)
        ~3 MB / 20 sheets. We extract the two muni/estado annexes:
          --sheet=muni    → 'Detalle por municipio' (76 cols × 2,472 munis)
          --sheet=estado  → 'Anexo-Estado'          (72 cols × 32 estados)

Why Python: \copy can't read XLSX (binary). Existing project pattern
(coneval-ageb-xlsx-to-csv.py, aeropuertos-xlsx-to-csv.py) — openpyxl with
data_only=True is the proven path.

Output schemas (snake_case ASCII, all data emitted as TEXT for raw table):

  --sheet=muni  (76 cols + 5 IDs):
    clave_municipio_num, cve_mun, nom_ent, nom_mun, nom_ent_mun,
    poblacion_total, poblacion_adulta, rezago_social,
    sucursales_{bm,bd,socap,sofipo,total},
    corresponsales_max,
    cajeros_{bm,bd,socap,sofipo,total},
    tpv_{bm,bd,socap,sofipo,total_eacp,agregadores,adq_no_banc,total_ag_adq,total},
    puntos_acceso_sca,
    cuentas_{bm,bd,socap,sofipo,total},
    creditos_{bm,bd,socap,sofipo,total},
    tx_tpv_{bm,bd,socap,sofipo,total},
    remesas_mdd, remesas_per_capita,
    g_cuentas_{bm,bd,socap,sofipo,total}_{m,h,b},
    g_creditos_{bm,bd,socap,sofipo,total}_{m,h,b}

  --sheet=estado (72 effective cols, last empty col dropped):
    cve_estado_num, nom_ent, poblacion_total, poblacion_adulta,
    sucursales_*, corresponsales_max, cajeros_*, tpv_*,
    cuentas_*, creditos_*,
    sar_{asignado,registrado,total},
    seg_{vida,pensiones,accidentes,danos_sin_autos,automoviles,total},
    tx_tpv_*,
    remesas_mdd,
    condusef_{ubicacion,reclamaciones},
    ac_inf_{sucursales,corresponsales,cajeros,tpv,total_ag_adq,estado},
    ac_pf_{captacion,credito,afore,vida,pensiones,accidentes,danos_sin_autos,automoviles,estado},
    ac_mp_{tx_tpv,remesas,estado_a,ubicacion,reclamaciones,estado_b}

Filtering:
  - Skip rows above data start (muni: row<5, estado: row<7).
  - Skip footer note rows (cell 1 starts with 'Nota:' or '*' or is None where
    cell 2 is also None).
  - Exclude catch-all sentinel row (clave_municipio_num=99999 / cve_estado_num=99
    "No identificado").

Sentinels:
  - Brecha cells (-999): mapped to '*' (CNBV's n<100 floor — see .xlsx note row).
  - Empty / None cells: mapped to '*'.
  - The downstream view uses NULLIF(col, '*') ::numeric to surface NULLs.
  - This converges "missing" and "n<100 suppressed" to a single sentinel —
    consumers can't distinguish the two. Documented in scan-cnbv-panorama-2025.md.

Usage:
  python3 cnbv-panorama-xlsx-to-csv.py --sheet=muni \
      raw/cnbv/Anexo_Panorama_2025.xlsx > /tmp/cnbv_panorama_muni.csv
  python3 cnbv-panorama-xlsx-to-csv.py --sheet=estado \
      raw/cnbv/Anexo_Panorama_2025.xlsx > /tmp/cnbv_panorama_estado.csv
"""

import csv
import sys

import openpyxl


MUNI_HEADERS = [
    "clave_municipio_num",
    "cve_mun",
    "nom_ent",
    "nom_mun",
    "nom_ent_mun",
    "poblacion_total",
    "poblacion_adulta",
    "rezago_social",
    # Sucursales (5)
    "sucursales_bm",
    "sucursales_bd",
    "sucursales_socap",
    "sucursales_sofipo",
    "sucursales_total",
    # Corresponsales (1)
    "corresponsales_max",
    # Cajeros (5)
    "cajeros_bm",
    "cajeros_bd",
    "cajeros_socap",
    "cajeros_sofipo",
    "cajeros_total",
    # TPV (9)
    "tpv_bm",
    "tpv_bd",
    "tpv_socap",
    "tpv_sofipo",
    "tpv_total_eacp",
    "tpv_agregadores",
    "tpv_adq_no_banc",
    "tpv_total_ag_adq",
    "tpv_total",
    # Punto acceso (1)
    "puntos_acceso_sca",
    # Cuentas (5)
    "cuentas_bm",
    "cuentas_bd",
    "cuentas_socap",
    "cuentas_sofipo",
    "cuentas_total",
    # Créditos (5)
    "creditos_bm",
    "creditos_bd",
    "creditos_socap",
    "creditos_sofipo",
    "creditos_total",
    # Tx TPV (5)
    "tx_tpv_bm",
    "tx_tpv_bd",
    "tx_tpv_socap",
    "tx_tpv_sofipo",
    "tx_tpv_total",
    # Remesas (2)
    "remesas_mdd",
    "remesas_per_capita",
    # Brechas Cuentas (15)
    "g_cuentas_bm_m",
    "g_cuentas_bm_h",
    "g_cuentas_bm_b",
    "g_cuentas_bd_m",
    "g_cuentas_bd_h",
    "g_cuentas_bd_b",
    "g_cuentas_socap_m",
    "g_cuentas_socap_h",
    "g_cuentas_socap_b",
    "g_cuentas_sofipo_m",
    "g_cuentas_sofipo_h",
    "g_cuentas_sofipo_b",
    "g_cuentas_total_m",
    "g_cuentas_total_h",
    "g_cuentas_total_b",
    # Brechas Créditos (15)
    "g_creditos_bm_m",
    "g_creditos_bm_h",
    "g_creditos_bm_b",
    "g_creditos_bd_m",
    "g_creditos_bd_h",
    "g_creditos_bd_b",
    "g_creditos_socap_m",
    "g_creditos_socap_h",
    "g_creditos_socap_b",
    "g_creditos_sofipo_m",
    "g_creditos_sofipo_h",
    "g_creditos_sofipo_b",
    "g_creditos_total_m",
    "g_creditos_total_h",
    "g_creditos_total_b",
]
assert len(MUNI_HEADERS) == 76, f"muni headers count {len(MUNI_HEADERS)} != 76"

ESTADO_HEADERS = [
    "cve_estado_num",
    "nom_ent",
    "poblacion_total",
    "poblacion_adulta",
    # Sucursales (5)
    "sucursales_bm",
    "sucursales_bd",
    "sucursales_socap",
    "sucursales_sofipo",
    "sucursales_total",
    # Corresponsales (1)
    "corresponsales_max",
    # Cajeros (5)
    "cajeros_bm",
    "cajeros_bd",
    "cajeros_socap",
    "cajeros_sofipo",
    "cajeros_total",
    # TPV (9)
    "tpv_bm",
    "tpv_bd",
    "tpv_socap",
    "tpv_sofipo",
    "tpv_total_eacp",
    "tpv_agregadores",
    "tpv_adq_no_banc",
    "tpv_total_ag_adq",
    "tpv_total",
    # Cuentas (5)
    "cuentas_bm",
    "cuentas_bd",
    "cuentas_socap",
    "cuentas_sofipo",
    "cuentas_total",
    # Créditos (5)
    "creditos_bm",
    "creditos_bd",
    "creditos_socap",
    "creditos_sofipo",
    "creditos_total",
    # SAR (3)
    "sar_asignado",
    "sar_registrado",
    "sar_total",
    # Seguros (6)
    "seg_vida",
    "seg_pensiones",
    "seg_accidentes",
    "seg_danos_sin_autos",
    "seg_automoviles",
    "seg_total",
    # Tx TPV (5)
    "tx_tpv_bm",
    "tx_tpv_bd",
    "tx_tpv_socap",
    "tx_tpv_sofipo",
    "tx_tpv_total",
    # Remesas (1 — estado sheet has only mdd, no per_capita)
    "remesas_mdd",
    # CONDUSEF (2)
    "condusef_ubicacion",
    "condusef_reclamaciones",
    # Acomodo - Infraestructura (6 ranking cols)
    "ac_inf_sucursales",
    "ac_inf_corresponsales",
    "ac_inf_cajeros",
    "ac_inf_tpv",
    "ac_inf_total_ag_adq",
    "ac_inf_estado",
    # Acomodo - productos financieros (9 ranking cols)
    "ac_pf_captacion",
    "ac_pf_credito",
    "ac_pf_afore",
    "ac_pf_vida",
    "ac_pf_pensiones",
    "ac_pf_accidentes",
    "ac_pf_danos_sin_autos",
    "ac_pf_automoviles",
    "ac_pf_estado",
    # Acomodo - medios de pago (6 ranking cols)
    "ac_mp_tx_tpv",
    "ac_mp_remesas",
    "ac_mp_estado_a",
    "ac_mp_ubicacion",
    "ac_mp_reclamaciones",
    "ac_mp_estado_b",
]
assert (
    len(ESTADO_HEADERS) == 72
), f"estado headers count {len(ESTADO_HEADERS)} != 72"


def cell_to_text(val):
    """Normalize a cell for downstream `\\copy NULL '*'` ingestion.

    None / empty / whitespace-only → '*' (which the view's NULLIF unwraps).
    Brecha sentinel -999 → '*' (CNBV's n<100 statistical-validity floor).
    Numeric values pass through (pandas csv writer handles the formatting).
    """
    if val is None:
        return "*"
    if isinstance(val, str):
        s = val.strip()
        return s if s else "*"
    if val == -999:
        return "*"
    return val


def emit_muni(ws, out):
    """Sheet shape:
    Row 1: title, Row 2: nav, Row 3: section banner, Row 4: column headers,
    Row 5+: data, Row 2478+: footer notes.
    Catch-all: clave_municipio_num=99999 'No identificado'.
    """
    out.writerow(MUNI_HEADERS)
    n_kept = 0
    n_skipped = 0
    for row_num in range(5, ws.max_row + 1):
        # Use index access to support 76 columns deterministically.
        clave_num = ws.cell(row_num, 1).value
        # Skip blank rows + footer note rows (clave_num is None or string starting
        # with text like 'Nota:', '*', '**').
        if clave_num is None:
            n_skipped += 1
            continue
        if isinstance(clave_num, str):
            n_skipped += 1
            continue
        # Catch-all sentinel — exclude
        if clave_num == 99999:
            n_skipped += 1
            continue
        out_row = [cell_to_text(ws.cell(row_num, c).value) for c in range(1, 77)]
        out.writerow(out_row)
        n_kept += 1
    sys.stderr.write(
        f"cnbv-panorama-xlsx-to-csv [muni]: kept={n_kept} skipped={n_skipped}\n"
    )


def emit_estado(ws, out):
    """Sheet shape:
    Row 1-2: title/nav, Row 3: numeric IDs, Row 4: nat'l totals, Row 5: section,
    Row 6: column headers, Row 7+: data, Row ~40+: footer notes.
    Catch-all: cve_estado_num=99 'No identificado'.
    Drop col 73 (always None).
    """
    out.writerow(ESTADO_HEADERS)
    n_kept = 0
    n_skipped = 0
    for row_num in range(7, ws.max_row + 1):
        clave_num = ws.cell(row_num, 1).value
        if clave_num is None or isinstance(clave_num, str):
            n_skipped += 1
            continue
        if clave_num == 99:
            # Catch-all sentinel — exclude
            n_skipped += 1
            continue
        out_row = [cell_to_text(ws.cell(row_num, c).value) for c in range(1, 73)]
        out.writerow(out_row)
        n_kept += 1
    sys.stderr.write(
        f"cnbv-panorama-xlsx-to-csv [estado]: kept={n_kept} skipped={n_skipped}\n"
    )


def main():
    sheet = None
    path = None
    for arg in sys.argv[1:]:
        if arg.startswith("--sheet="):
            sheet = arg.split("=", 1)[1]
        elif not arg.startswith("--"):
            path = arg
    if sheet not in ("muni", "estado") or not path:
        sys.stderr.write(
            "usage: cnbv-panorama-xlsx-to-csv.py --sheet={muni|estado} <path-to-xlsx>\n"
        )
        sys.exit(2)

    wb = openpyxl.load_workbook(path, read_only=False, data_only=True)
    sheet_name = "Detalle por municipio" if sheet == "muni" else "Anexo-Estado"
    if sheet_name not in wb.sheetnames:
        sys.stderr.write(
            f"error: expected sheet {sheet_name!r}, got {wb.sheetnames}\n"
        )
        sys.exit(3)
    ws = wb[sheet_name]

    out = csv.writer(sys.stdout)
    if sheet == "muni":
        emit_muni(ws, out)
    else:
        emit_estado(ws, out)


if __name__ == "__main__":
    main()
