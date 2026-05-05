#!/usr/bin/env python3
r"""
Convert CONEVAL GRS_AGEB_urbana_2020.xlsx to a clean CSV on stdout.

Source: https://www.coneval.org.mx/Medicion/Documents/GRS_AGEB_2020/
        GRS_AGEB_urbana_2020.zip → GRS_AGEB_urbana_2020.xlsx (61,444 rows × 28 cols)

Why Python: \copy can't read XLSX (binary), and there's no maintained
Node/TS library for streaming-read of an Excel workbook. openpyxl with
read_only + data_only handles the 16 MB file in ~12s with constant memory.

Output schema (21 columns, all valid AGEB rows):
  cvegeo, pobtot, vivpar_hab,
  ind_analfabeta, ind_no_escuela_6_14, ind_no_escuela_15_24,
  ind_basica_incompleta, ind_sin_salud, ind_hacinamiento,
  ind_sin_agua, ind_sin_excusado, ind_sin_drenaje, ind_sin_luz,
  ind_piso_tierra, ind_sin_lavadora, ind_sin_refri, ind_sin_telfijo,
  ind_sin_celular, ind_sin_compu, ind_sin_internet,
  grado

Filtering:
  - Skip header rows 1-6 (workbook has multi-row title + section headers).
  - Skip footer "Nota:" rows (last ~6 rows have a single text cell at col 0).
  - Skip rows where col 7 (Clave de la AGEB) is not 13-char A-Z0-9.
  - Skip rows where col 27 (Grado) is null/empty (footer / orphan rows).

INEGI/CONEVAL data caveats:
  - Per the data dictionary, AGEBs with <3 viviendas habitadas have ALL
    indicators NULL but `grado` populated (LSNIEG art. 37 confidentiality).
    We pass these through — the loader stores NULL for indicators and
    keeps the grado.
  - Empty cells are emitted as empty strings; \copy NULL '' converts.

Usage:
  python3 coneval-ageb-xlsx-to-csv.py /path/to/GRS_AGEB_urbana_2020.xlsx > /tmp/coneval_grs.csv
"""

import csv
import re
import sys

import openpyxl


HEADER = [
    "cvegeo",
    "pobtot",
    "vivpar_hab",
    "ind_analfabeta",
    "ind_no_escuela_6_14",
    "ind_no_escuela_15_24",
    "ind_basica_incompleta",
    "ind_sin_salud",
    "ind_hacinamiento",
    "ind_sin_agua",
    "ind_sin_excusado",
    "ind_sin_drenaje",
    "ind_sin_luz",
    "ind_piso_tierra",
    "ind_sin_lavadora",
    "ind_sin_refri",
    "ind_sin_telfijo",
    "ind_sin_celular",
    "ind_sin_compu",
    "ind_sin_internet",
    "grado",
]

# 12 digits + 1 alphanumeric (letter suffix permitted, ~9% of AGEBs).
CVEGEO_RE = re.compile(r"^[0-9]{12}[0-9A-Z]$")

VALID_GRADOS = {"Muy bajo", "Bajo", "Medio", "Alto", "Muy alto"}


def cell(val):
    """Normalize a cell value for downstream `\\copy NULL '*'` ingestion.

    Returns `*` for None / empty / whitespace-only values so the loader's
    NULL '*' directive maps them to actual NULL. CONEVAL itself uses `*`
    as the LSNIEG art. 37 confidentiality sentinel; this just collapses
    "missing" and "suppressed" into the same value, which the view's
    `NULLIF(col, '*')::numeric` casts handle uniformly. Without this,
    a None cell would emit `""` to CSV → load as empty TEXT in the raw
    table → `NULLIF('', '*')` returns `""` → `''::numeric` throws at
    view SELECT time. qa-audit C1 (2026-05-05).
    """
    if val is None:
        return "*"
    if isinstance(val, str):
        s = val.strip()
        return s if s else "*"
    return val


def main():
    if len(sys.argv) != 2:
        sys.stderr.write(
            "usage: coneval-ageb-xlsx-to-csv.py <path-to-GRS_AGEB_urbana_2020.xlsx>\n",
        )
        sys.exit(2)
    path = sys.argv[1]

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "GRS 2020" not in wb.sheetnames:
        sys.stderr.write(
            f"error: expected sheet 'GRS 2020', got {wb.sheetnames}\n",
        )
        sys.exit(3)
    ws = wb["GRS 2020"]

    out = csv.writer(sys.stdout)
    out.writerow(HEADER)

    n_kept = 0
    n_skipped = 0
    for row in ws.iter_rows(min_row=7, values_only=True):
        # Defensive: some workbooks have trailing all-None rows beyond max_row.
        if row is None:
            n_skipped += 1
            continue
        cvegeo = cell(row[7]) if len(row) > 7 else ""
        if not isinstance(cvegeo, str) or not CVEGEO_RE.match(cvegeo):
            n_skipped += 1
            continue
        grado = cell(row[27]) if len(row) > 27 else ""
        if grado not in VALID_GRADOS:
            # Skip rows with a malformed/missing grado — the dataset's contract
            # is that every AGEB has at least the Grado populated, so a row
            # missing it is a footer artifact or schema corruption.
            n_skipped += 1
            continue
        out.writerow(
            [
                cvegeo,
                cell(row[8]),  # Población total
                cell(row[9]),  # Viviendas particulares habitadas
                cell(row[10]),  # 17 indicators in fixed order...
                cell(row[11]),
                cell(row[12]),
                cell(row[13]),
                cell(row[14]),
                cell(row[15]),
                cell(row[16]),
                cell(row[17]),
                cell(row[18]),
                cell(row[19]),
                cell(row[20]),
                cell(row[21]),
                cell(row[22]),
                cell(row[23]),
                cell(row[24]),
                cell(row[25]),
                cell(row[26]),
                grado,
            ]
        )
        n_kept += 1

    sys.stderr.write(f"coneval-ageb-xlsx-to-csv: kept={n_kept} skipped={n_skipped}\n")


if __name__ == "__main__":
    main()
