#!/usr/bin/env python3
"""
Convert SCT/AFAC airport pivot XLSX to a flat CSV ready for the TS loader.

The source XLSX (gob.mx producto-aeropuertos-2006-2026-mar-NNNNNNNN.xlsx) has a
pivot table on sheet "TD Prod Aptos" where:
  - Operator codes appear as parent rows: AICM, AIFA, ASA, ASUR, GACM, OMA,
    GAP, MIDCM, CHIH-PAC, TOL (and a few longer codes like CHIHUAHUA, CENTRO
    NORTE, PACIFICO).
  - City/airport names are child rows beneath their operator.
  - The same airport (e.g. CD. DEL CARMEN) may appear under multiple
    operators historically when ownership transferred.

Output CSV shape (header row + one row per (airport_name, operator, ano)):
  airport_name, operator, ano, mar_flights

Usage:
  python3 scripts/aeropuertos-xlsx-to-csv.py \\
    raw/airports/producto-aeropuertos.xlsx \\
    raw/airports/producto-aeropuertos.csv

Notes:
  - We track the current "operator" as we walk down the rows. The operator
    label is short (≤8 chars), all-uppercase letters/spaces/hyphens, and
    appears alone (not duplicated as a city name on subsequent rows in this
    sheet — the duplicated CIUDAD DE MÉXICO/SANTA LUCÍA cases are a 1:1
    operator-with-one-airport pattern, which we DO want to keep as the city
    name).
  - The downstream view dedupes across (airport_name, ano) so the same
    physical airport appearing under two operators (e.g. POZA RICA listed
    under ASA legacy and GAP current) is summed into one row per year.
"""

import csv
import re
import sys
from openpyxl import load_workbook

# Pattern for operator-parent rows. ALL-UPPERCASE codes ≤8 chars, letters
# only (allow hyphens for CHIH-PAC). Note that some "city names" also fit
# this pattern (e.g. AICM = "Mexico City") but those have data in their row
# AND are followed by a second row with the longer city name. We keep the
# operator label and skip duplicates downstream via (airport_name, operator,
# ano) primary key.
KNOWN_OPERATORS = {
    "AICM", "AIFA", "ASA", "ASUR", "GACM", "OMA", "GAP", "MIDCM",
    "CHIH-PAC", "CHIHUAHUA", "TOL", "TOLUCA", "MEXICANA",
    "CENTRO NORTE", "PACIFICO", "AMP",
}

# Years in the column header (R5: 2006..2026)
START_YEAR = 2006
END_YEAR = 2026


def main() -> int:
    if len(sys.argv) < 3:
        print("Usage: aeropuertos-xlsx-to-csv.py <input.xlsx> <output.csv>", file=sys.stderr)
        return 1
    xlsx = sys.argv[1]
    out_csv = sys.argv[2]

    wb = load_workbook(xlsx, data_only=True, read_only=True)
    if "TD Prod Aptos" not in wb.sheetnames:
        print(f"ERROR: sheet 'TD Prod Aptos' missing from {xlsx}", file=sys.stderr)
        return 2
    ws = wb["TD Prod Aptos"]

    # Read year header row (R5). Columns 2..22 are 2006..2026.
    rows = list(ws.iter_rows(min_row=5, max_row=5, values_only=True))
    if not rows:
        print("ERROR: empty header row", file=sys.stderr)
        return 2
    hdr = rows[0]
    years = []
    for v in hdr[1:]:
        try:
            y = int(v) if v is not None else None
        except (TypeError, ValueError):
            y = None
        years.append(y)
    if not all(y is not None for y in years[:21]):
        print(f"ERROR: header years parse failed: {years}", file=sys.stderr)
        return 2

    operator_now = "UNKNOWN"
    n_airport_rows = 0
    n_operator_rows = 0
    out_rows: list[tuple[str, str, int, int]] = []

    for row in ws.iter_rows(min_row=6, values_only=True):
        if not row or row[0] is None:
            continue
        label = str(row[0]).strip()
        if not label:
            continue
        # Skip totals
        if label.lower() in ("total", "total general", "grand total"):
            continue
        # Operator detection
        if label in KNOWN_OPERATORS:
            operator_now = label
            n_operator_rows += 1
            continue
        # Otherwise — airport row. Extract per-year values.
        n_airport_rows += 1
        for i, year in enumerate(years[:21], start=0):
            val = row[1 + i]
            if val is None:
                continue
            try:
                flights = int(val)
            except (TypeError, ValueError):
                try:
                    flights = int(float(val))
                except (TypeError, ValueError):
                    continue
            # Drop zero-rows entirely — they bloat the table without info.
            if flights == 0:
                continue
            out_rows.append((label, operator_now, year, flights))

    # Write CSV
    with open(out_csv, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f, quoting=csv.QUOTE_MINIMAL)
        w.writerow(["airport_name", "operator", "ano", "mar_flights"])
        w.writerows(out_rows)

    # Stats
    distinct_airports = len({r[0] for r in out_rows})
    print(
        f"[aeropuertos-xlsx-to-csv] {n_airport_rows} airport rows + {n_operator_rows} operator rows seen",
        file=sys.stderr,
    )
    print(
        f"[aeropuertos-xlsx-to-csv] {len(out_rows)} non-zero (airport,operator,year) cells, "
        f"{distinct_airports} distinct airports → {out_csv}",
        file=sys.stderr,
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
